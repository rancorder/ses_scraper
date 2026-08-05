"""
storage/storage.py - 結果の保存・読み込み
  ・CSV / JSON / Excel 出力
  ・Excel はスコア順・候補ハイライト付き
"""
from __future__ import annotations

import csv
import json
import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from company_analyzer.config import OUTPUT_DIR
from company_analyzer.models import ScoredCompany

log = logging.getLogger(__name__)


def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def save_json(results: list[ScoredCompany], path: Path | None = None) -> Path:
    path = path or OUTPUT_DIR / f"results_{_timestamp()}.json"
    data = [r.model_dump() for r in results]
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    log.info(f"JSON保存 → {path}")
    return path


def save_csv(results: list[ScoredCompany], path: Path | None = None) -> Path:
    path = path or OUTPUT_DIR / f"results_{_timestamp()}.csv"
    if not results:
        log.warning("保存するデータがありません")
        return path

    rows = []
    for r in results:
        rows.append({
            "会社名":         r.company_name,
            "URL":            r.company_url,
            "ドメイン":       r.domain,
            "スコア":         r.score,
            "営業候補":       "◎" if r.is_candidate else "",
            "自社製品":       "○" if r.score_breakdown.get("product_presence") else "",
            "OEM対応":        "○" if r.score_breakdown.get("oem_presence") else "",
            "技術キーワード": "|".join(r.tech_keywords.keys()),
            "検出特徴":       " / ".join(r.detected_features),
            "製品ページ":     " ".join(r.product_pages[:3]),
            "OEMページ":      " ".join(r.oem_pages[:3]),
            "メール":         r.contact_email or "",
            "フォームあり":   "○" if r.has_contact_form else "",
            "取得ページ数":   r.pages_crawled,
            "エラー":         r.error or "",
        })

    df = pd.DataFrame(rows)
    df.to_csv(path, index=False, encoding="utf-8-sig")
    log.info(f"CSV保存 → {path}  ({len(results)}社)")
    return path


def save_excel(results: list[ScoredCompany], path: Path | None = None) -> Path:
    path = path or OUTPUT_DIR / f"results_{_timestamp()}.xlsx"
    if not results:
        log.warning("保存するデータがありません")
        return path

    rows = []
    for r in results:
        breakdown_str = " / ".join(
            f"{k}:{v}" for k, v in r.score_breakdown.items() if v > 0
        )
        rows.append({
            "会社名":         r.company_name,
            "URL":            r.company_url,
            "スコア":         r.score,
            "営業候補":       "◎" if r.is_candidate else "",
            "自社製品":       "○" if r.score_breakdown.get("product_presence") else "",
            "OEM対応":        "○" if r.score_breakdown.get("oem_presence") else "",
            "技術カテゴリ":   "|".join(r.tech_keywords.keys()),
            "検出特徴":       " / ".join(r.detected_features[:5]),
            "製品ページ":     r.product_pages[0] if r.product_pages else "",
            "OEMページ":      r.oem_pages[0] if r.oem_pages else "",
            "メール":         r.contact_email or "",
            "フォームあり":   "○" if r.has_contact_form else "",
            "取得ページ数":   r.pages_crawled,
            "スコア内訳":     breakdown_str,
            "エラー":         r.error or "",
        })

    df = pd.DataFrame(rows)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="スコアリング結果")
        ws = writer.sheets["スコアリング結果"]

        # ヘッダースタイル
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", name="Arial")
        center      = Alignment(horizontal="center", vertical="center")
        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = center
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

        # 営業候補行をハイライト
        cand_fill    = PatternFill("solid", fgColor="E2EFDA")   # 薄緑
        high_fill    = PatternFill("solid", fgColor="FFF2CC")   # 薄黄
        score_col    = df.columns.get_loc("スコア") + 1
        cand_col     = df.columns.get_loc("営業候補") + 1

        for row_idx in range(2, ws.max_row + 1):
            score_cell = ws.cell(row=row_idx, column=score_col)
            score = int(score_cell.value or 0)
            is_candidate = ws.cell(row=row_idx, column=cand_col).value == "◎"

            if is_candidate:
                for col in range(1, len(df.columns) + 1):
                    ws.cell(row=row_idx, column=col).fill = cand_fill
            elif score >= 50:
                for col in range(1, len(df.columns) + 1):
                    ws.cell(row=row_idx, column=col).fill = high_fill

        # URLハイパーリンク
        link_font = Font(color="0563C1", underline="single", name="Arial")
        url_col   = df.columns.get_loc("URL") + 1
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=url_col)
            url = str(cell.value or "").strip()
            if url.startswith("http"):
                cell.hyperlink = url
                cell.font = link_font

        # 列幅
        widths = {
            "会社名": 28, "URL": 35, "スコア": 8, "営業候補": 8,
            "自社製品": 8, "OEM対応": 8, "技術カテゴリ": 25,
            "検出特徴": 45, "製品ページ": 30, "OEMページ": 30,
            "メール": 28, "フォームあり": 10, "取得ページ数": 12,
            "スコア内訳": 40, "エラー": 20,
        }
        for cell in ws[1]:
            ws.column_dimensions[cell.column_letter].width = widths.get(str(cell.value), 15)

        # サマリーシート
        candidates = [r for r in results if r.is_candidate]
        _add_summary_sheet(writer, results, candidates)

    log.info(f"Excel保存 → {path}  (全{len(results)}社 / 候補{len([r for r in results if r.is_candidate])}社)")
    return path


def _add_summary_sheet(writer, all_results, candidates):
    """サマリーシートを追加"""
    summary_data = {
        "項目": [
            "解析企業数", "営業候補数（70点以上）",
            "候補率", "平均スコア",
            "自社製品保有数", "OEM対応数",
            "IoT関連数", "組み込み関連数",
        ],
        "値": [
            len(all_results),
            len(candidates),
            f"{len(candidates)/len(all_results)*100:.1f}%" if all_results else "0%",
            f"{sum(r.score for r in all_results)/len(all_results):.1f}" if all_results else "0",
            sum(1 for r in all_results if r.score_breakdown.get("product_presence")),
            sum(1 for r in all_results if r.score_breakdown.get("oem_presence")),
            sum(1 for r in all_results if "iot" in r.tech_keywords),
            sum(1 for r in all_results if "embedded" in r.tech_keywords),
        ],
    }
    pd.DataFrame(summary_data).to_excel(writer, index=False, sheet_name="サマリー")


def load_companies_from_excel(path: str | Path) -> list[dict]:
    """
    IPROSスクレイパー出力Excelから企業リストを読み込む。
    「会社名」「公式サイト」「詳細URL」列を使用。
    """
    path = Path(path)
    all_companies = []

    sheets = pd.read_excel(path, sheet_name=None, dtype=str)
    for sheet_name, df in sheets.items():
        df = df.fillna("")
        for _, row in df.iterrows():
            url = str(row.get("公式サイト", "")).strip()
            if not url or not url.startswith("http"):
                # 公式サイトがなければIPROSの詳細URLにフォールバック
                url = str(row.get("詳細URL", "")).strip()
            if not url or not url.startswith("http"):
                continue
            all_companies.append({
                "name":    str(row.get("会社名", url)).strip(),
                "url":     url,
                "keyword": str(row.get("検索キーワード", "")).strip(),
            })

    # URLで重複除去
    seen = set()
    unique = []
    for c in all_companies:
        if c["url"] not in seen:
            seen.add(c["url"])
            unique.append(c)

    log.info(f"企業読み込み: {len(unique)} 社 ({path.name})")
    return unique
