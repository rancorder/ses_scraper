"""
merge_results.py - バッチ出力Excelを1ファイルに統合
"""
import argparse, sys
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, '/opt/ses_scraper/SES_scra_anaraiz')

_FILL_EXCELLENT = PatternFill("solid", fgColor="C6EFCE")
_FILL_GOOD      = PatternFill("solid", fgColor="FFEB9C")
_FILL_FAIR      = PatternFill("solid", fgColor="FFCCCC")
_FILL_HEADER    = PatternFill("solid", fgColor="2F5496")
_FONT_HEADER    = Font(color="FFFFFF", bold=True)
_LINK_FONT      = Font(color="0563C1", underline="single")

def merge(prefix, output, output_dir=None):
    from ses_config import OUTPUT_DIR
    out_dir = Path(output_dir) if output_dir else OUTPUT_DIR
    files = sorted(out_dir.glob(f"*{prefix}*batch*.xlsx"))
    if not files:
        files = sorted(out_dir.glob("*batch*.xlsx"))
    if not files:
        print(f"❌ 対象ファイルが見つかりません")
        return

    print(f"対象ファイル数: {len(files)}")
    all_dfs = []
    for f in files:
        try:
            df = pd.read_excel(f, sheet_name=0, dtype=str).fillna("")
            all_dfs.append(df)
            print(f"  ✓ {f.name}: {len(df)}行")
        except Exception as e:
            print(f"  ❌ {f.name}: {e}")

    if not all_dfs:
        print("❌ 読み込めるファイルがありませんでした")
        return

    df_all = pd.concat(all_dfs, ignore_index=True)
    print(f"\n統合後: {len(df_all)}行")

    score_col = next((c for c in df_all.columns if "スコア" in c and "AI" not in c), None)
    if score_col:
        df_all[score_col] = pd.to_numeric(df_all[score_col], errors="coerce").fillna(0)
        df_all = df_all.sort_values(score_col, ascending=False).reset_index(drop=True)

    output_path = out_dir / output
    wb = Workbook()
    ws = wb.active
    ws.title = "統合結果"
    headers = list(df_all.columns)
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    judgment_idx = next((i+1 for i, h in enumerate(headers) if "判定" in str(h)), None)
    url_idx      = next((i+1 for i, h in enumerate(headers) if h in ["サイトURL","URL","企業ホームページURL","公式サイト"]), None)

    for _, row in df_all.iterrows():
        ws.append(list(row))
        row_idx  = ws.max_row
        judgment = str(row.iloc[judgment_idx-1] if judgment_idx else "")
        fill = {"◎": _FILL_EXCELLENT, "○": _FILL_GOOD, "△": _FILL_FAIR}.get(judgment)
        if fill:
            for col in range(1, len(headers)+1):
                ws.cell(row=row_idx, column=col).fill = fill

    if url_idx:
        for row_idx in range(2, ws.max_row+1):
            cell = ws.cell(row=row_idx, column=url_idx)
            val = str(cell.value or "").strip()
            if val.startswith("http"):
                cell.hyperlink = val
                cell.font = _LINK_FONT

    for i, col in enumerate(headers, 1):
        w = 28 if "会社名" in col or "企業名" in col else 38 if "URL" in col else 8 if "判定" in col else 12 if "スコア" in col else 18
        ws.column_dimensions[get_column_letter(i)].width = w

    ws2 = wb.create_sheet("サマリー")
    total     = len(df_all)
    j_col     = headers[judgment_idx-1] if judgment_idx else None
    excellent = int((df_all[j_col] == "◎").sum()) if j_col else 0
    good      = int((df_all[j_col] == "○").sum()) if j_col else 0
    fair      = int((df_all[j_col] == "△").sum()) if j_col else 0

    for row in [
        ["項目", "値"],
        ["統合ファイル数", len(files)],
        ["総企業数",       total],
        ["◎ 優先候補",     excellent],
        ["○ 候補",         good],
        ["△ 要確認",       fair],
        ["候補合計",       excellent+good],
        ["候補率",         f"{(excellent+good)/total*100:.1f}%" if total else "0%"],
    ]:
        ws2.append(row)

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 15
    for cell in ws2[1]:
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER

    wb.save(output_path)
    print(f"\n✅ 統合完了: {output_path}")
    print(f"   ◎{excellent}社 / ○{good}社 / △{fair}社 / 計{total}社")

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--prefix",     default="")
    parser.add_argument("--output",     default="統合結果.xlsx")
    parser.add_argument("--output-dir", default=None)
    args = parser.parse_args()
    merge(args.prefix, args.output, args.output_dir)
