"""
scraper/salesnow.py - SalesNow URL指定スクレイパー＋リアルタイム評価
====================================================================
URLを指定して企業リストを収集し、1社ずつ company_analyzer で即時評価。
70点以上の候補のみ Excel にリアルタイム追記。
"""
from __future__ import annotations

import asyncio
import logging
import random
import re
import sys
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Optional
from urllib.parse import urljoin

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from playwright.async_api import async_playwright, BrowserContext, Page
from playwright.async_api import TimeoutError as PWTimeout

log = logging.getLogger(__name__)

BASE_URL   = "https://salesnow.jp"
DELAY_MIN  = 1.5
DELAY_MAX  = 3.0

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/122.0.0.0 Safari/537.36"
)


# ─── データモデル ──────────────────────────────────────────────────────
@dataclass
class SalesNowCompany:
    ソース:          str = "SalesNow"
    会社名:          str = ""
    法人番号:        str = ""
    説明:            str = ""
    SalesNowスコア:  str = ""
    従業員数:        str = ""
    資本金:          str = ""
    所在地:          str = ""
    公式サイト:      str = ""
    詳細URL:         str = ""

    def to_dict(self) -> dict:
        return asdict(self)


# ─── HTML解析 ──────────────────────────────────────────────────────────

def _parse_company_list(html: str) -> list[dict]:
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html, "lxml")
    stubs = []
    seen = set()
    for a in soup.select("a[href*='/db/companies/']"):
        href = a.get("href", "")
        name = a.get_text(strip=True)
        if not href or not name:
            continue
        full = urljoin(BASE_URL, href)
        if full in seen:
            continue
        seen.add(full)
        stubs.append({"name": name, "detail_url": full})
    return stubs


def _parse_next_page(html: str, current_url: str) -> Optional[str]:
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html, "lxml")
    for a in soup.select("a[href*='/db/search/page/'], a[href*='/db/industries/']"):
        txt = a.get_text(strip=True)
        if txt in ("次へ", "›", ">", "NEXT", "次のページ"):
            return urljoin(BASE_URL, a["href"])
    links = soup.select("a[href*='/db/search/page/']")
    hrefs = [urljoin(BASE_URL, a["href"]) for a in links]
    if current_url in hrefs:
        idx = hrefs.index(current_url)
        if idx + 1 < len(hrefs):
            return hrefs[idx + 1]
    return None


def _parse_detail(html: str, detail_url: str) -> SalesNowCompany:
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html, "lxml")
    c = SalesNowCompany(詳細URL=detail_url)
    body = soup.get_text(" ", strip=True)

    h1 = soup.select_one("h1")
    if h1:
        c.会社名 = h1.get_text(strip=True)

    m = re.search(r"法人番号[：:]\s*(\d{13})", body)
    if m: c.法人番号 = m.group(1)

    m = re.search(r"SalesNowスコア\s*([A-Z][+\-]?評価)", body)
    if m: c.SalesNowスコア = m.group(1)

    m = re.search(r"従業員数\s*[\n\s]*([\d,]+)\s*名", body)
    if m: c.従業員数 = m.group(1) + "名"

    m = re.search(r"資本金\s*[\n\s]*([\d,]+\s*万円)", body)
    if m: c.資本金 = m.group(1)

    m = re.search(r"(東京都|大阪府|京都府|北海道|.{2,3}県)[^\s　]{1,30}[区市町村]", body)
    if m: c.所在地 = m.group(0)[:30]

    for sel in ["[class*='description'] p", "section p", ".summary p"]:
        t = soup.select_one(sel)
        if t and len(t.get_text(strip=True)) > 20:
            c.説明 = t.get_text(strip=True)[:300]
            break

    for a in soup.select("a[href^='http']"):
        href = a["href"]
        if ("salesnow.jp" not in href and "wantedly" not in href
                and "google" not in href and len(href) > 10):
            c.公式サイト = href
            break

    return c


# ─── Playwright フェッチ ────────────────────────────────────────────────

async def _fetch(page: Page, url: str, timeout_ms: int = 15000) -> Optional[str]:
    try:
        resp = await page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
        if resp is None or resp.status in (404, 403, 410):
            return None
        try:
            await page.wait_for_load_state("networkidle", timeout=2500)
        except PWTimeout:
            pass
        return await page.content()
    except PWTimeout:
        return None
    except Exception as e:
        log.debug(f"  取得エラー ({type(e).__name__}): {url}")
        return None


# ─── リアルタイムExcel追記 ─────────────────────────────────────────────

EXCEL_COLUMNS = [
    "会社名", "スコア", "営業候補", "公式サイト", "ソース",
    "SalesNowスコア", "従業員数", "資本金", "所在地", "法人番号",
    "自社製品", "OEM対応", "技術キーワード", "検出特徴",
    "メール", "フォーム", "取得ページ数", "SalesNow詳細URL",
]


def _init_excel(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "営業候補"
    ws.append(EXCEL_COLUMNS)
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF", name="游ゴシック")
        cell.fill      = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    widths = {"会社名": 28, "スコア": 8, "営業候補": 10, "公式サイト": 40, "ソース": 12,
              "SalesNowスコア": 14, "従業員数": 10, "資本金": 12, "所在地": 25,
              "技術キーワード": 30, "検出特徴": 40}
    for col_idx, col_name in enumerate(EXCEL_COLUMNS, 1):
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = widths.get(col_name, 15)
    wb.save(path)


def _append_candidate(path: Path, company: SalesNowCompany, scored) -> None:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    row_num = ws.max_row + 1

    fill_color = "E2EFDA" if scored.score >= 70 else "FFF2CC"
    fill = PatternFill("solid", fgColor=fill_color)
    link_font   = Font(color="0563C1", underline="single", name="游ゴシック")
    normal_font = Font(name="游ゴシック")

    values = {
        "会社名":          company.会社名,
        "スコア":          scored.score,
        "営業候補":        "◎" if scored.is_candidate else "△",
        "公式サイト":      company.公式サイト,
        "ソース":          company.ソース,
        "SalesNowスコア":  company.SalesNowスコア,
        "従業員数":        company.従業員数,
        "資本金":          company.資本金,
        "所在地":          company.所在地,
        "法人番号":        company.法人番号,
        "自社製品":        "○" if scored.score_breakdown.get("product_presence") else "",
        "OEM対応":         "○" if scored.score_breakdown.get("oem_presence") else "",
        "技術キーワード":  "|".join(scored.tech_keywords.keys()),
        "検出特徴":        " / ".join(scored.detected_features[:5]),
        "メール":          scored.contact_email or "",
        "フォーム":        "○" if scored.has_contact_form else "",
        "取得ページ数":    scored.pages_crawled,
        "SalesNow詳細URL": company.詳細URL,
    }

    for col_idx, col_name in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.value = values.get(col_name, "")
        cell.fill  = fill
        cell.font  = normal_font
        if col_name == "公式サイト" and str(values.get("公式サイト", "")).startswith("http"):
            cell.hyperlink = values["公式サイト"]
            cell.font = link_font
        elif col_name == "SalesNow詳細URL" and str(values.get("SalesNow詳細URL", "")).startswith("http"):
            cell.hyperlink = values["SalesNow詳細URL"]
            cell.font = link_font

    wb.save(path)


# ─── 1社評価 ──────────────────────────────────────────────────────────

async def _evaluate_one(company: SalesNowCompany, context: BrowserContext):
    url = company.公式サイト
    if not url or not url.startswith("http"):
        return None

    _root = Path(__file__).resolve().parent.parent
    if str(_root) not in sys.path:
        sys.path.insert(0, str(_root))

    try:
        from company_analyzer.crawler.crawler import crawl_site
        from company_analyzer.parser.site_parser import parse_page
        from company_analyzer.feature_extractor.extractor import extract_features
        from company_analyzer.scoring.scoring_engine import ScoringEngine
    except ImportError as e:
        log.error(f"  company_analyzer インポートエラー: {e}")
        return None

    sem = asyncio.Semaphore(1)
    pages = await crawl_site(url, context, sem)
    if not pages:
        return None

    parsed = [parse_page(p.url, p.html) for p in pages if p.html]
    from urllib.parse import urlparse
    domain   = urlparse(url).netloc
    features = extract_features(domain, parsed)
    engine   = ScoringEngine()
    scored   = engine.score_company(company.会社名, url, features)
    return scored


# ─── メイン評価ループ ──────────────────────────────────────────────────

async def scrape_and_evaluate_async(
    start_urls: list[str],
    output: str = "salesnow_candidates.xlsx",
    threshold: int = 70,
    max_pages: int = 10,
    stop_event: Optional[asyncio.Event] = None,
) -> dict:
    output_path = Path(output)
    _init_excel(output_path)
    log.info(f"  出力先: {output_path} (閾値:{threshold}点)")

    stats = {"total": 0, "evaluated": 0, "candidates": 0, "no_url": 0}
    seen_detail: set[str] = set()

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-blink-features=AutomationControlled", "--disable-dev-shm-usage"],
        )
        context: BrowserContext = await browser.new_context(
            user_agent=USER_AGENT, locale="ja-JP", timezone_id="Asia/Tokyo",
            ignore_https_errors=True,
            extra_http_headers={"Accept-Language": "ja,en-US;q=0.9,en;q=0.8"},
        )
        await context.add_init_script(
            "Object.defineProperty(navigator, 'webdriver', { get: () => undefined });"
        )
        try:
            for start_url in start_urls:
                if stop_event and stop_event.is_set():
                    break

                page_url = start_url
                page_num = 0
                log.info(f"\n[SalesNow] URL: {start_url}")

                list_page = await context.new_page()
                try:
                    while page_url and page_num < max_pages:
                        if stop_event and stop_event.is_set():
                            break
                        page_num += 1
                        log.info(f"  一覧 p{page_num}: {page_url}")
                        html = await _fetch(list_page, page_url)
                        if not html:
                            break

                        stubs = _parse_company_list(html)
                        if not stubs:
                            log.info("  企業リストなし → 終了")
                            break

                        log.info(f"  {len(stubs)} 社を取得 → 各社を即時評価")

                        for stub in stubs:
                            if stop_event and stop_event.is_set():
                                break
                            if stub["detail_url"] in seen_detail:
                                continue
                            seen_detail.add(stub["detail_url"])
                            stats["total"] += 1

                            await asyncio.sleep(random.uniform(DELAY_MIN, DELAY_MAX))
                            detail_html = await _fetch(list_page, stub["detail_url"])
                            if not detail_html:
                                company = SalesNowCompany(会社名=stub["name"], 詳細URL=stub["detail_url"])
                            else:
                                company = _parse_detail(detail_html, stub["detail_url"])
                                if not company.会社名:
                                    company.会社名 = stub["name"]

                            if not company.公式サイト:
                                stats["no_url"] += 1
                                log.info(f"    ⬜ {company.会社名[:30]:<30} 公式URLなし → スキップ")
                                continue

                            scored = await _evaluate_one(company, context)
                            if not scored:
                                stats["no_url"] += 1
                                log.info(f"    ⬜ {company.会社名[:30]:<30} 評価失敗")
                                continue

                            stats["evaluated"] += 1
                            mark = "◎" if scored.score >= threshold else "－"
                            log.info(
                                f"    {mark} [{scored.score:3d}点] {company.会社名[:30]:<30}"
                                f" | {company.公式サイト[:35]}"
                            )

                            if scored.score >= threshold:
                                stats["candidates"] += 1
                                _append_candidate(output_path, company, scored)
                                log.info(f"      → ✅ 候補追記 ({stats['candidates']}社目)")

                        next_url = _parse_next_page(html, page_url)
                        page_url = next_url if (next_url and next_url != page_url) else None
                        if page_url:
                            await asyncio.sleep(random.uniform(1.0, 2.0))
                finally:
                    await list_page.close()

        finally:
            await context.close()
            await browser.close()

    log.info(
        f"\n[SalesNow] 完了: 収集{stats['total']}社 / "
        f"評価{stats['evaluated']}社 / "
        f"候補{stats['candidates']}社 / "
        f"URL無{stats['no_url']}社"
    )
    return stats


def scrape_and_evaluate(
    start_urls: list[str],
    output: str = "salesnow_candidates.xlsx",
    threshold: int = 70,
    max_pages: int = 10,
    stop_event=None,
) -> dict:
    """同期ラッパー"""
    async def _run():
        async_stop = asyncio.Event()
        if stop_event:
            async def _watch():
                while not async_stop.is_set():
                    if stop_event.is_set():
                        async_stop.set()
                    await asyncio.sleep(0.5)
            asyncio.create_task(_watch())
        return await scrape_and_evaluate_async(
            start_urls, output, threshold, max_pages,
            async_stop if stop_event else None
        )
    return asyncio.run(_run())
