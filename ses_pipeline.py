"""
ses_pipeline.py - SES事業スクリーニング専用パイプライン
=======================================================
フロー:
  入力（Excel/IPROS/SalesNow）
    ↓ [1] クロール（既存 crawler.py 流用）
    ↓ [2] テキスト抽出
    ↓ [3] キーワード一次フィルタ（ses_config のキーワード辞書）
    ↓ [4] Ollama AI判定（スコア・理由・営業トーク）
    ↓ [5] Excel出力（発注側スコア・供給側スコア・理由・営業トーク列付き）

変更履歴:
  - 住所・電話・ソース・検索キーワード列をExcel出力に追加
  - source_path 指定時は元ファイルに評価列を追記する形式で出力
"""
from __future__ import annotations

import asyncio
import logging
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional
from urllib.parse import urlparse

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

# パスを通す
_HERE = Path(__file__).resolve().parent
if str(_HERE) not in sys.path:
    sys.path.insert(0, str(_HERE))


# ════════════════════════════════════════════════════════
#  データモデル
# ════════════════════════════════════════════════════════

@dataclass
class SesResult:
    """1社分のSESスクリーニング結果"""
    company_name:    str
    company_url:     str

    # ── IPROS等から引き継ぐ基本情報 ──────────────────────
    住所:            str = ""
    電話:            str = ""
    検索キーワード:  str = ""
    ソース:          str = ""

    # キーワード一次スコア
    kw_client_score:  int = 0
    kw_partner_score: int = 0
    kw_hits_client:   list[str] = field(default_factory=list)
    kw_hits_partner:  list[str] = field(default_factory=list)

    # AI判定スコア（Ollama）
    ai_client_score:  int = 0
    ai_partner_score: int = 0
    ai_reason:        str = ""
    ai_sales_talk:    str = ""
    ai_error:         Optional[str] = None

    # 総合
    final_client_score:  int = 0
    final_partner_score: int = 0
    judgment:            str = ""   # ◎/○/△/－

    # プロファイルスコアリング結果
    profile_score: object = None   # type: ignore

    pages_crawled: int = 0
    error:         Optional[str] = None

    def calc_final_scores(self, profile=None) -> None:
        if self.profile_score is not None:
            ps = self.profile_score
            self.final_client_score  = ps.total_score
            self.final_partner_score = min(100, int(self.ai_partner_score * 0.7)) if self.ai_partner_score else 0
            self.judgment = ps.judgment
            if self.ai_client_score:
                ai_bonus = int(self.ai_client_score * 0.15)
                self.final_client_score = min(100, ps.total_score + ai_bonus)
                th_exc = getattr(profile, "threshold_excellent", 70) if profile else 70
                th_gd  = getattr(profile, "threshold_good", 50) if profile else 50
                if self.final_client_score >= th_exc:
                    self.judgment = "◎"
                elif self.final_client_score >= th_gd:
                    self.judgment = "○"
                elif self.final_client_score > 0:
                    self.judgment = "△"
                else:
                    self.judgment = "－"
        else:
            kw_c = min(self.kw_client_score * 5, 30)
            kw_p = min(self.kw_partner_score * 5, 30)
            self.final_client_score  = min(100, kw_c + int(self.ai_client_score  * 0.7))
            self.final_partner_score = min(100, kw_p + int(self.ai_partner_score * 0.7))
            best = max(self.final_client_score, self.final_partner_score)
            th_exc = getattr(profile, "threshold_excellent", 70) if profile else 70
            th_gd  = getattr(profile, "threshold_good", 50) if profile else 50
            if best >= th_exc:
                self.judgment = "◎"
            elif best >= th_gd:
                self.judgment = "○"
            elif best >= 30:
                self.judgment = "△"
            else:
                self.judgment = "－"


# ════════════════════════════════════════════════════════
#  Step 1: Excel / CSV からの企業リスト読み込み
# ════════════════════════════════════════════════════════

def load_companies(path: str | Path) -> list[dict]:
    path = Path(path)
    if path.suffix.lower() in (".xlsx", ".xls"):
        sheets = pd.read_excel(path, sheet_name=None, dtype=str)
        rows = []
        for df in sheets.values():
            rows.extend(df.fillna("").to_dict(orient="records"))
    elif path.suffix.lower() == ".csv":
        rows = pd.read_csv(path, dtype=str).fillna("").to_dict(orient="records")
    else:
        raise ValueError(f"未対応ファイル形式: {path.suffix}")

    results = []
    for row in rows:
        name = _find_col(row, ["会社名", "企業名", "company_name", "name", "社名"])
        url  = _find_col(row, ["公式サイト", "URL", "url", "ホームページ", "website", "サイト"])
        if not name and not url:
            continue
        if url and not url.startswith("http"):
            url = "https://" + url
        results.append({"name": name or url, "url": url or ""})

    seen, unique = set(), []
    for c in results:
        key = c["url"] or c["name"]
        if key not in seen:
            seen.add(key)
            unique.append(c)

    log.info(f"  読み込み: {len(unique)} 社 ({path.name})")
    return unique


def _find_col(row: dict, candidates: list[str]) -> str:
    for key in candidates:
        val = str(row.get(key, "")).strip()
        if val:
            return val
    return ""


# ════════════════════════════════════════════════════════
#  Step 2: クロール
# ════════════════════════════════════════════════════════

async def _crawl(companies: list[dict], concurrency: int, crawl_paths: list[str] | None = None) -> dict[str, list]:
    import asyncio, time
    from company_analyzer.crawler.crawler import crawl_all

    total = len(companies)
    safe_concurrency = min(concurrency, 4)
    log.info(f"  対象: {total}社 / 同時接続: {safe_concurrency}（VPS安定化）")

    start = time.monotonic()
    async def _hb():
        interval = 0
        while True:
            await asyncio.sleep(30)
            interval += 30
            elapsed_min = interval // 60
            log.info(f"  ⏳ クロール継続中... ({elapsed_min}分{interval%60}秒経過)")
    hb = asyncio.create_task(_hb())

    try:
        timeout_sec = min(int(total / safe_concurrency * 30 * 2), 10800)
        log.info(f"  タイムアウト設定: {timeout_sec//60}分")
        results = await asyncio.wait_for(
            crawl_all(companies, safe_concurrency, paths=crawl_paths),
            timeout=timeout_sec,
        )
    except asyncio.TimeoutError:
        log.warning(f"  ⚠ クロールタイムアウト ({timeout_sec//60}分) → 取得済み分で継続")
        results = {}
    finally:
        hb.cancel()

    ok = sum(1 for v in results.values() if v)
    elapsed = int(time.monotonic() - start)
    log.info(f"  クロール完了: {ok}/{total}社取得 ({elapsed}秒)")
    return results


# ════════════════════════════════════════════════════════
#  Step 3: テキスト抽出
# ════════════════════════════════════════════════════════

def _extract_text(pages: list) -> str:
    parts = []
    for p in pages:
        if hasattr(p, "text") and p.text:
            parts.append(p.text)
        elif hasattr(p, "title") and p.title:
            parts.append(p.title)
    return "\n".join(parts)


# ════════════════════════════════════════════════════════
#  Step 4: キーワード一次フィルタ
# ════════════════════════════════════════════════════════

def _keyword_scan(
    text: str,
    client_keywords: dict | None = None,
    partner_keywords: dict | None = None,
) -> tuple[list[str], list[str]]:
    if client_keywords is None or partner_keywords is None:
        from ses_config import CLIENT_KEYWORDS, PARTNER_KEYWORDS
        client_keywords  = client_keywords  or CLIENT_KEYWORDS
        partner_keywords = partner_keywords or PARTNER_KEYWORDS

    text_lower = text.lower()
    client_hits  = []
    partner_hits = []

    for category, words in client_keywords.items():
        for w in words:
            if w.lower() in text_lower:
                client_hits.append(w)

    for category, words in partner_keywords.items():
        for w in words:
            if w.lower() in text_lower:
                partner_hits.append(w)

    return client_hits, partner_hits


# ════════════════════════════════════════════════════════
#  Step 5: Excel出力
# ════════════════════════════════════════════════════════

_FILL_EXCELLENT = PatternFill("solid", fgColor="C6EFCE")
_FILL_GOOD      = PatternFill("solid", fgColor="FFEB9C")
_FILL_FAIR      = PatternFill("solid", fgColor="FFCCCC")
_FILL_HEADER    = PatternFill("solid", fgColor="2F5496")
_FONT_HEADER    = Font(color="FFFFFF", bold=True)


def _build_result_row(r: SesResult, axes_defs: list, profile=None) -> list:
    """SesResult から評価列のリストを生成する（共通処理）"""
    # 評価軸スコア列
    if axes_defs and r.profile_score:
        ax_results = {ax.id: ax for ax in r.profile_score.axes}
        axis_cols = []
        for ax_def in axes_defs:
            ax_res = ax_results.get(ax_def.get("id", ""))
            if ax_res:
                if ax_res.earned:
                    cell_val = f"✓ {ax_res.score}点"
                    if ax_res.hit_keywords:
                        cell_val += "\n(" + ", ".join(ax_res.hit_keywords[:2]) + ")"
                else:
                    cell_val = "✗ 0点"
            else:
                cell_val = "-"
            axis_cols.append(cell_val)
    else:
        axis_cols = [
            r.final_client_score,
            r.final_partner_score,
            ", ".join(r.kw_hits_client[:6]),
            ", ".join(r.kw_hits_partner[:6]),
        ]

    ai_cols   = [r.ai_client_score or "", r.ai_reason or "", r.ai_sales_talk or ""]
    tail_cols = [r.pages_crawled, r.error or ""]

    return axis_cols + ai_cols + tail_cols


def save_ses_excel(results: list[SesResult], output_path: Path, profile=None) -> None:
    """結果を新規Excelに保存（従来形式）"""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws = wb.active
    profile_name = profile.name if profile else "SES"
    ws.title = f"{profile_name[:20]}_結果"

    axes_defs    = (profile.scoring_axes if profile and profile.scoring_axes else [])
    axis_headers = [f"{ax['name']}({ax['points']}点)" for ax in axes_defs]

    base_headers  = ["判定", "総合スコア", "会社名", "サイトURL", "電話", "住所", "ソース", "検索KW"]
    score_headers = axis_headers if axis_headers else ["発注側スコア", "供給側スコア", "発注側KWヒット", "供給側KWヒット"]
    ai_headers    = ["AI補正スコア", "AI判定理由", "営業トーク案"]
    tail_headers  = ["クロールページ数", "エラー"]
    headers       = base_headers + score_headers + ai_headers + tail_headers

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sorted_results = sorted(results, key=lambda x: x.final_client_score, reverse=True)
    for r in sorted_results:
        base_cols = [
            r.judgment, r.final_client_score, r.company_name, r.company_url,
            r.電話, r.住所, r.ソース, r.検索キーワード,
        ]
        row_data = base_cols + _build_result_row(r, axes_defs, profile)
        ws.append(row_data)

        row_idx = ws.max_row
        fill = {"◎": _FILL_EXCELLENT, "○": _FILL_GOOD, "△": _FILL_FAIR}.get(r.judgment)
        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = fill

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 38
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 18

    score_start = 9
    for i in range(score_start, score_start + len(score_headers)):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = 22

    ai_start = score_start + len(score_headers)
    for i, w in enumerate([12, 50, 60], ai_start):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    link_font = Font(color="0563C1", underline="single")
    url_col_idx = headers.index("サイトURL") + 1
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=url_col_idx)
        url_val = str(cell.value or "").strip()
        if url_val.startswith("http"):
            cell.hyperlink = url_val
            cell.font = link_font

    for row in ws.iter_rows(min_row=2, min_col=score_start, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    _add_summary_and_axes_sheets(wb, results, profile)

    wb.save(output_path)
    log.info(f"  保存完了: {output_path}")


def save_ses_excel_append(
    results: list[SesResult],
    source_path: Path,
    output_path: Path,
    profile=None,
    url_col: str = "URL",
) -> None:
    """
    元ファイルの列をそのまま残して、右側に評価列を追記して保存する。
    URLをキーに結果を紐付ける。
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # ── 元ファイル読み込み ────────────────────────────────
    src = Path(source_path)
    if src.suffix.lower() == ".csv":
        for enc in ["utf-8", "cp932", "shift_jis"]:
            try:
                df_src = pd.read_csv(src, dtype=str, encoding=enc).fillna("")
                break
            except Exception:
                continue
        else:
            df_src = pd.read_csv(src, dtype=str, encoding="utf-8", errors="replace").fillna("")
    else:
        df_src = pd.read_excel(src, dtype=str).fillna("")

    # ── 結果をURLをキーに辞書化 ───────────────────────────
    result_map: dict[str, SesResult] = {}
    for r in results:
        key = r.company_url.rstrip("/")
        result_map[key] = r

    # ── 評価軸ヘッダー生成 ────────────────────────────────
    axes_defs    = (profile.scoring_axes if profile and profile.scoring_axes else [])
    axis_headers = [f"{ax['name']}({ax['points']}点)" for ax in axes_defs]
    score_headers = axis_headers if axis_headers else ["発注側スコア", "供給側スコア", "発注側KWヒット", "供給側KWヒット"]

    append_headers = (
        ["【判定】", "【総合スコア】"] +
        [f"【{h}】" for h in score_headers] +
        ["【AI補正スコア】", "【AI判定理由】", "【営業トーク案】",
         "【クロールページ数】", "【エラー】"]
    )

    # ── 追記列をDataFrameとして生成 ──────────────────────
    append_rows = []
    for _, row in df_src.iterrows():
        url_val = str(row.get(url_col, "") or "").strip().rstrip("/")
        # URL列が見つからない場合は他の候補も探す
        if not url_val:
            for col_candidate in ["公式サイト", "URL", "url", "ホームページ", "website"]:
                url_val = str(row.get(col_candidate, "") or "").strip().rstrip("/")
                if url_val:
                    break

        r = result_map.get(url_val)
        if r:
            score_cols = _build_result_row(r, axes_defs, profile)
            append_rows.append([r.judgment, r.final_client_score] + score_cols)
        else:
            # マッチしなかった行は空欄
            append_rows.append([""] * len(append_headers))

    df_append = pd.DataFrame(append_rows, columns=append_headers)
    df_out = pd.concat([df_src.reset_index(drop=True), df_append], axis=1)

    # ── Excel書き出し ────────────────────────────────────
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = Workbook()
    ws = wb.active
    profile_name = profile.name if profile else "SES"
    ws.title = f"{profile_name[:20]}_結果"

    # ヘッダー行
    all_headers = list(df_out.columns)
    ws.append(all_headers)

    orig_col_count = len(df_src.columns)
    append_col_start = orig_col_count + 1  # 追記列の開始列番号

    # 元列ヘッダー：グレー
    orig_header_fill = PatternFill("solid", fgColor="D9D9D9")
    orig_header_font = Font(bold=True, color="333333")
    # 追記列ヘッダー：紺
    append_header_fill = PatternFill("solid", fgColor="2F5496")
    append_header_font = Font(color="FFFFFF", bold=True)

    for i, cell in enumerate(ws[1], 1):
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if i <= orig_col_count:
            cell.fill = orig_header_fill
            cell.font = orig_header_font
        else:
            cell.fill = append_header_fill
            cell.font = append_header_font
    ws.row_dimensions[1].height = 22

    # データ行
    for _, row in df_out.iterrows():
        ws.append(list(row))

        row_idx = ws.max_row
        judgment = str(row.get("【判定】", "") or "")
        fill = {"◎": _FILL_EXCELLENT, "○": _FILL_GOOD, "△": _FILL_FAIR}.get(judgment)
        if fill:
            # 追記列のみハイライト
            for col in range(append_col_start, len(all_headers) + 1):
                ws.cell(row=row_idx, column=col).fill = fill

    # 列幅（元列は自動、追記列は固定）
    for i in range(1, orig_col_count + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    append_widths = [8, 12] + [22] * len(score_headers) + [12, 50, 60, 12, 15]
    for i, w in enumerate(append_widths, append_col_start):
        ws.column_dimensions[get_column_letter(i)].width = w

    # URLハイパーリンク（元列のURL列を探して設定）
    link_font = Font(color="0563C1", underline="single")
    url_col_names = ["URL", "公式サイト", "url", "ホームページ", "website", "サイト"]
    url_col_idx = None
    for col_name in url_col_names:
        if col_name in all_headers:
            url_col_idx = all_headers.index(col_name) + 1
            break
    if url_col_idx:
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=url_col_idx)
            url_val = str(cell.value or "").strip()
            if url_val.startswith("http"):
                cell.hyperlink = url_val
                cell.font = link_font

    # 追記列の折り返し
    for row in ws.iter_rows(min_row=2, min_col=append_col_start, max_col=len(all_headers)):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"

    # サマリー・評価軸別集計シートを追加
    _add_summary_and_axes_sheets(wb, results, profile)

    wb.save(output_path)
    log.info(f"  追記保存完了: {output_path}  (元{orig_col_count}列 + 追記{len(append_headers)}列)")


def _add_summary_and_axes_sheets(wb, results: list[SesResult], profile=None) -> None:
    """サマリーシートと評価軸別集計シートを追加（共通処理）"""
    axes_defs = (profile.scoring_axes if profile and profile.scoring_axes else [])
    profile_name = profile.name if profile else "SES"
    total     = len(results)
    excellent = sum(1 for r in results if r.judgment == "◎")
    good      = sum(1 for r in results if r.judgment == "○")
    fair      = sum(1 for r in results if r.judgment == "△")
    th_exc    = getattr(profile, "threshold_excellent", 70) if profile else 70
    th_gd     = getattr(profile, "threshold_good",      50) if profile else 50

    ws2 = wb.create_sheet("サマリー")
    summary_rows = [
        ["項目", "値"],
        ["案件プロファイル",              profile_name],
        ["解析企業数",                    total],
        [f"◎ 優先候補（{th_exc}点以上）", excellent],
        [f"○ 候補（{th_gd}点以上）",      good],
        ["△ 要確認",                      fair],
        ["候補合計",                      excellent + good],
        ["候補率",                        f"{(excellent+good)/total*100:.1f}%" if total else "0%"],
        ["AI判定成功数",                  sum(1 for r in results if r.ai_client_score)],
        ["クロール失敗数",                sum(1 for r in results if r.error)],
        ["平均スコア",                    f"{sum(r.final_client_score for r in results)/total:.1f}" if total else "0"],
    ]
    for s_row in summary_rows:
        ws2.append(s_row)
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 20
    for cell in ws2[1]:
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER

    if axes_defs:
        ws3 = wb.create_sheet("評価軸別集計")
        ws3.append(["評価軸", "配点", "ヒット社数", "ヒット率", "主なヒットKW（上位3）"])
        for cell in ws3[1]:
            cell.fill = _FILL_HEADER
            cell.font = _FONT_HEADER

        for ax_def in axes_defs:
            ax_id  = ax_def.get("id", "")
            ax_nm  = ax_def.get("name", "")
            ax_pts = ax_def.get("points", 0)
            hits   = []
            kw_counter: dict[str, int] = {}
            for r in results:
                if r.profile_score:
                    ax_res = next((a for a in r.profile_score.axes if a.id == ax_id), None)
                    if ax_res and ax_res.earned:
                        hits.append(r.company_name)
                        for kw in ax_res.hit_keywords:
                            kw_counter[kw] = kw_counter.get(kw, 0) + 1
            top_kws  = sorted(kw_counter, key=lambda k: -kw_counter[k])[:3]
            hit_rate = f"{len(hits)/total*100:.1f}%" if total else "0%"
            ws3.append([ax_nm, f"{ax_pts}点", len(hits), hit_rate, "、".join(top_kws)])

        for i, w in enumerate([30, 10, 12, 10, 40], 1):
            ws3.column_dimensions[ws3.cell(1, i).column_letter].width = w


# ════════════════════════════════════════════════════════
#  メインパイプライン
# ════════════════════════════════════════════════════════

async def run_ses_pipeline(
    companies: list[dict],
    output_prefix: str = "ses_analysis",
    concurrency: int = 4,
    use_ollama: bool = True,
    profile=None,
    source_path: str | Path | None = None,  # ← 追加: 元ファイルパス
    url_col: str = "URL",                   # ← 追加: 元ファイルのURL列名
) -> list[SesResult]:
    """
    SESスクリーニングのメインパイプライン。

    Args:
        companies:    [{"name": ..., "url": ..., "住所": ..., "電話": ..., ...}, ...]
        source_path:  元のCSV/Excelパス。指定時は元ファイルに評価列を追記して出力。
        url_col:      source_path 内のURL列名（デフォルト: "URL"）
    """
    from ses_config import OUTPUT_DIR, SES_SCORE_CFG
    from company_analyzer.parser.site_parser import parse_page
    from ollama_scorer import OllamaScorer, create_scorer

    profile_name = profile.name if profile else "SES（デフォルト）"
    client_kws   = profile.client_keywords  if profile else None
    partner_kws  = profile.partner_keywords if profile else None
    if profile:
        output_prefix = f"{profile.slug}_{output_prefix.lstrip('ses_')}"

    total = len(companies)
    log.info("=" * 60)
    log.info(f"スクリーニング開始: {total}社  案件: [{profile_name}]")
    log.info("=" * 60)

    scorer = None
    if use_ollama:
        scorer = create_scorer()
        if profile and profile.ai_instruction:
            scorer.custom_instruction = profile.ai_instruction
        if scorer.is_available():
            models = scorer.get_available_models()
            log.info(f"  Ollama 接続OK | モデル: {scorer.model}")
            log.info(f"  利用可能モデル: {', '.join(models[:5])}")
        else:
            log.warning("  ⚠ Ollamaに接続できません → キーワード判定のみで実行します")
            scorer = None

    log.info(f"\n[1/4] クロール開始 ({total}社 / 同時{concurrency}接続)")
    # プロファイルのcrawl_pathsがあればそちらを優先
    _crawl_paths = getattr(profile, "crawl_paths", None) if profile else None
    crawl_results = await _crawl(companies, concurrency, crawl_paths=_crawl_paths)

    log.info("\n[2-3/4] テキスト抽出 & キーワードスキャン中...")
    results: list[SesResult] = []

    for i, c in enumerate(companies, 1):
        url   = c.get("url", "")
        name  = c.get("name", url)
        pages = crawl_results.get(url, [])

        result = SesResult(
            company_name   = name,
            company_url    = url,
            住所           = c.get("住所", "") or "",
            電話           = c.get("電話", "") or "",
            検索キーワード = c.get("keyword", "") or "",
            ソース         = c.get("source", "") or "",
        )
        result.pages_crawled = len(pages)

        if not pages:
            result.error = "no_pages"
            results.append(result)
        else:
            parsed = [parse_page(p.url, p.html) for p in pages if p.html]
            site_text = _extract_text(parsed)

            ch, ph = _keyword_scan(site_text, client_kws, partner_kws)
            result.kw_hits_client  = ch
            result.kw_hits_partner = ph
            result.kw_client_score  = len(ch)
            result.kw_partner_score = len(ph)
            result._site_text = site_text  # type: ignore[attr-defined]

            if profile and getattr(profile, "scoring_axes", []):
                try:
                    from profile_scorer import ProfileScorer
                    p_scorer = ProfileScorer(profile)
                    result.profile_score = p_scorer.score(name, url, parsed)
                except Exception as e_ps:
                    log.debug(f"  ProfileScorer エラー ({name}): {e_ps}")

            results.append(result)

        if i % 20 == 0 or i == total:
            hit_c = sum(1 for r in results if r.kw_client_score > 0)
            hit_p = sum(1 for r in results if r.kw_partner_score > 0)
            no_pg = sum(1 for r in results if r.error == "no_pages")
            log.info(f"  KWスキャン {i}/{total}社 | 発注ヒット:{hit_c} 供給ヒット:{hit_p} クロール失敗:{no_pg}")

    log.info(f"  完了: {len(results)}社")

    if scorer:
        OLLAMA_PARALLEL = 3
        ollama_sem = asyncio.Semaphore(OLLAMA_PARALLEL)
        targets = [r for r in results if not r.error and getattr(r, "_site_text", "").strip()]

        log.info(f"\n[4/4] Ollama AI判定中... ({scorer.model}) [{len(targets)}社 / 並列{OLLAMA_PARALLEL}]")
        completed = 0

        async def _score_one(r: SesResult) -> None:
            nonlocal completed
            async with ollama_sem:
                site_text = getattr(r, "_site_text", "")
                loop = asyncio.get_event_loop()
                ollama_result = await loop.run_in_executor(
                    None, scorer.score, r.company_name, site_text
                )
                r.ai_client_score  = ollama_result.client_score
                r.ai_partner_score = ollama_result.partner_score
                r.ai_reason        = ollama_result.reason
                r.ai_sales_talk    = ollama_result.sales_talk
                r.ai_error         = ollama_result.error
                completed += 1
                verdict_mark = "◎" if r.ai_client_score >= 60 or r.ai_partner_score >= 60 else "  "
                log.info(
                    f"  [{completed:>3}/{len(targets)}] {verdict_mark} "
                    f"発注:{r.ai_client_score:>3} 供給:{r.ai_partner_score:>3} "
                    f"({ollama_result.elapsed_sec:.1f}秒) {r.company_name[:25]}"
                )

        await asyncio.gather(*[_score_one(r) for r in targets])
        log.info(f"  AI判定完了: {completed} 社")
    else:
        log.info("\n[4/4] AI判定スキップ（キーワード判定のみ）")

    log.info("\n最終スコア計算 & 保存中...")
    for r in results:
        r.calc_final_scores(profile=profile)

    candidates = [r for r in results if r.judgment in ("◎", "○")]
    log.info(f"  営業候補: ◎{sum(1 for r in results if r.judgment=='◎')}社 "
             f"○{sum(1 for r in results if r.judgment=='○')}社 "
             f"（計{len(candidates)}社 / {total}社中）")

    # ── 保存（source_path があれば追記形式、なければ従来形式）──
    output_path = OUTPUT_DIR / f"{output_prefix}.xlsx"
    if source_path and Path(source_path).exists():
        log.info(f"  元ファイルに追記形式で保存: {source_path}")
        save_ses_excel_append(
            results     = results,
            source_path = Path(source_path),
            output_path = output_path,
            profile     = profile,
            url_col     = url_col,
        )
    else:
        save_ses_excel(results, output_path, profile=profile)

    log.info("\n" + "=" * 60)
    log.info("  🏆  TOP 10 SES営業候補")
    log.info("=" * 60)
    top = sorted(results, key=lambda x: max(x.final_client_score, x.final_partner_score), reverse=True)
    for i, r in enumerate(top[:10], 1):
        log.info(f"  {i:2}. [{r.judgment}] {r.company_name} | 発注:{r.final_client_score} 供給:{r.final_partner_score}")
        if r.ai_reason:
            log.info(f"      → {r.ai_reason[:60]}")

    return results
